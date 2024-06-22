import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Function to handle button click (process CSV and Excel)
def process_files():
    # Get the paths of CSV and Excel files
    csv_file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    if csv_file_path and excel_file_path:
        # Ask for the new sheet name
        new_sheet_name = entry_sheet_name.get()

        # Read CSV into pandas DataFrame
        df = pd.read_csv(csv_file_path)

        # Load existing Excel workbook
        wb = load_workbook(excel_file_path)

        # Create a new sheet and append data from DataFrame
        new_sheet = wb.create_sheet(title=new_sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(r)

        # Save the updated workbook
        wb.save(excel_file_path)
        status_label.config(text=f"Data copied to '{new_sheet_name}' in '{excel_file_path}'")

# Create tkinter window
window = tk.Tk()
window.title("CSV to Excel")

# Create and place UI elements
label_sheet_name = tk.Label(window, text="New Sheet Name:")
label_sheet_name.grid(row=0, column=0, padx=10, pady=10)

entry_sheet_name = tk.Entry(window)
entry_sheet_name.grid(row=0, column=1, padx=10, pady=10)

button_process = tk.Button(window, text="Process Files", command=process_files)
button_process.grid(row=1, columnspan=2, padx=10, pady=10)

status_label = tk.Label(window, text="")
status_label.grid(row=2, columnspan=2, padx=10, pady=10)

# Run the tkinter main loop
window.mainloop()
